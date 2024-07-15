#!/usr/bin/env python
# coding: utf-8


# # 세번째코드 (2024.7.12)

# ## 엑셀 파일 넣기

# In[18]:
import streamlit as st
import pandas as pd
import yfinance as yf

st.title("ETF Holdings 모니터링")

# Create an upload button
uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])

# In[ ]:
import pandas as pd



# In[ ]:
#if uploaded_file is not None:
#    # Read only column A from the file
#    df = pd.read_excel(uploaded_file, usecols="A")
#    st.write(df)



# In[ ]:
# Initialize a session state for dates if not already initialized
# Initialize a session state for dates if not already initialized
if 'dates' not in st.session_state:
    st.session_state.dates = [None, None, None]

# Function to add a new date input
def add_date():
    st.session_state.dates.append(None)

# Function to remove the last date input
def remove_date():
    if len(st.session_state.dates) > 0:
        st.session_state.dates.pop()

# Create the first date input widget for the end date
st.session_state.dates[0] = st.date_input("end date", key="date_input_0")

# Create date input widgets for the start dates
for i in range(1, len(st.session_state.dates), 3):
    cols = st.columns(3)
    for j in range(3):
        if i + j < len(st.session_state.dates):
            label = f"start date{i + j}" if i + j > 0 else "end date"
            with cols[j]:
                st.session_state.dates[i + j] = st.date_input(label, key=f"date_input_{i + j}")

# Display buttons to add or remove date inputs
st.button("Add Date", on_click=add_date)
st.button("Remove Date", on_click=remove_date)

def calculate_returns(df, end_date, start_dates):
    for i, ticker in enumerate(df['ticker']):
        try:
            ticker_data = yf.download(ticker, period='5y')
            end_price = ticker_data.loc[end_date, 'Close']
            
            for j, start_date in enumerate(start_dates):
                start_price = ticker_data.loc[start_date, 'Close']
                return_col = f'수익률{j+1}'
                if return_col not in df.columns:
                    df[return_col] = None
                df.loc[i, return_col] = (end_price / start_price - 1) * 100
        except Exception as e:
            print(f"Error fetching data for {ticker}: {e}")
    return df

if uploaded_file is not None:
    # Read the file
    df = pd.read_excel(uploaded_file)

    # Split the first column into 'ticker' and 'Country'
    new_columns = df.iloc[:, 0].str.split(' ', n=1, expand=True)
    new_columns.columns = ['ticker', 'Country']
    
    # Combine the new columns with the rest of the dataframe
    df = pd.concat([new_columns, df.iloc[:, 1:]], axis=1)
    
    # Convert the third and fourth columns to percentages
    if df.shape[1] > 2:
        df.iloc[:, 2] = df.iloc[:, 2].apply(lambda x: f"{x * 100:.2f}%")
    if df.shape[1] > 3:
        df.iloc[:, 3] = df.iloc[:, 3].apply(lambda x: f"{x * 100:.2f}%")
    
    # Display the dataframe before calculating returns
    st.write("Uploaded Data:", df)
    
    # Button to calculate returns
    if st.button("Check return rate"):
        # Convert date inputs to strings for compatibility with yfinance
        end_date = st.session_state.dates[0].strftime('%Y-%m-%d')
        start_dates = [date.strftime('%Y-%m-%d') for date in st.session_state.dates[1:] if date is not None]
        
        # Calculate returns and update the dataframe
        df = calculate_returns(df, end_date, start_dates)
        
        # Format the returns columns as percentages
        for col in df.columns:
            if '수익률' in col:
                df[col] = df[col].apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else None)
        
        # Display the modified dataframe
        st.write("Data with Returns:", df)



# In[ ]:
from openpyxl import Workbook

# Initialize a session state for dates if not already initialized
if 'dates' not in st.session_state:
    st.session_state.dates = [None, None, None]

# Function to add a new date input
def add_date():
    st.session_state.dates.append(None)

# Function to remove the last date input
def remove_date():
    if len(st.session_state.dates) > 0:
        st.session_state.dates.pop()

# Create the first date input widget for the end date
st.session_state.dates[0] = st.date_input("end date", key="date_input_0")

# Create date input widgets for the start dates
for i in range(1, len(st.session_state.dates), 3):
    cols = st.columns(3)
    for j in range(3):
        if i + j < len(st.session_state.dates):
            label = f"start date{i + j}" if i + j > 0 else "end date"
            with cols[j]:
                st.session_state.dates[i + j] = st.date_input(label, key=f"date_input_{i + j}")

# Display buttons to add or remove date inputs
st.button("Add Date", on_click=add_date)
st.button("Remove Date", on_click=remove_date)

def save_ticker_data_to_excel(ticker_data, filename="엑셀로 다운받기.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "종가raw"
    
    for idx, (ticker, data) in enumerate(ticker_data.items()):
        col = idx * 2 + 1
        ws.cell(row=1, column=col, value=ticker)
        ws.cell(row=1, column=col + 1, value="Date")
        for row_idx, (date, value) in enumerate(data['Close'].items(), start=2):
            ws.cell(row=row_idx, column=col, value=value)
            ws.cell(row=row_idx, column=col + 1, value=date.strftime('%Y-%m-%d'))

    wb.save(filename)

def fetch_ticker_data(tickers):
    ticker_data = {}
    for ticker in tickers:
        try:
            data = yf.download(ticker, period='5y')
            ticker_data[ticker] = data
        except Exception as e:
            print(f"Error fetching data for {ticker}: {e}")
    save_ticker_data_to_excel(ticker_data)
    return ticker_data

def calculate_returns(df, end_date, start_dates):
    for i, ticker in enumerate(df['ticker']):
        try:
            ticker_data = yf.download(ticker, period='5y')
            end_price = ticker_data.loc[end_date, 'Close']
            
            for j, start_date in enumerate(start_dates):
                start_price = ticker_data.loc[start_date, 'Close']
                return_col = f'수익률{j+1}'
                if return_col not in df.columns:
                    df[return_col] = None
                df.loc[i, return_col] = (end_price / start_price - 1) * 100
        except Exception as e:
            print(f"Error fetching data for {ticker}: {e}")
    return df

st.title("ETF Holdings 모니터링")

# Create an upload button
uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])

if uploaded_file is not None:
    # Read the file
    df = pd.read_excel(uploaded_file)

    # Split the first column into 'ticker' and 'Country'
    new_columns = df.iloc[:, 0].str.split(' ', n=1, expand=True)
    new_columns.columns = ['ticker', 'Country']
    
    # Combine the new columns with the rest of the dataframe
    df = pd.concat([new_columns, df.iloc[:, 1:]], axis=1)
    
    # Convert the third and fourth columns to percentages
    if df.shape[1] > 2:
        df.iloc[:, 2] = df.iloc[:, 2].apply(lambda x: f"{x * 100:.2f}%")
    if df.shape[1] > 3:
        df.iloc[:, 3] = df.iloc[:, 3].apply(lambda x: f"{x * 100:.2f}%")
    
    # Display the dataframe before calculating returns
    st.write("Uploaded Data:", df)
    
    # Button to calculate returns
    if st.button("Check return rate"):
        # Convert date inputs to strings for compatibility with yfinance
        end_date = st.session_state.dates[0].strftime('%Y-%m-%d')
        start_dates = [date.strftime('%Y-%m-%d') for date in st.session_state.dates[1:] if date is not None]
        
        # Fetch ticker data if not already fetched
        tickers = df['ticker'].tolist()
        ticker_data = fetch_ticker_data(tickers)
        
        # Calculate returns and update the dataframe
        df = calculate_returns(df, end_date, start_dates)
        
        # Format the returns columns as percentages
        for col in df.columns:
            if '수익률' in col:
                df[col] = df[col].apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else None)
        
        # Display the modified dataframe
        st.write("Data with Returns:", df)
        
        # Provide download link for the excel file
        with open("엑셀로 다운받기.xlsx", "rb") as file:
            st.download_button(
                label="Download Excel file",
                data=file,
                file_name="엑셀로 다운받기.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




